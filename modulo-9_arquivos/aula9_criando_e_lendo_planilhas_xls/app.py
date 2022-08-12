# Sheets = Páginas do arquivo XLS
# Workbook = Arquivo XLS

import os
import openpyxl

os.chdir('modulo-9_arquivos')
os.chdir('aula9_criando_e_lendo_planilhas_xls')

# Criar planilha
planilha_test = openpyxl.Workbook()
sheet_frutas = planilha_test.create_sheet('Frutas')
sheet_legumes = planilha_test.create_sheet('Legumes')
sheet_sementes = planilha_test.create_sheet('Sementes')
print(planilha_test.sheetnames)
cabecalho_frutas = ['Título', 'Localização', 'Preço']
sheet_frutas.append(cabecalho_frutas)

frutas = [['Uva', 'Mercado', 5], ['Melancia', 'Mercado', 15], ['Bolo', 'Mercado', 40]]
for fruta in frutas:
    sheet_frutas.append(fruta)

planilha_test.save(os.getcwd() + os.sep + 'Dados Supermercado.xlsx')
