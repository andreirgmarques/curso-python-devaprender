import os
import openpyxl

os.chdir('modulo-9_arquivos')
os.chdir('aula10_manipulando_planilhas_existentes')

caminho_planilha = os.getcwd() + os.sep + 'pessoas.xlsx'
planilha = openpyxl.load_workbook(caminho_planilha)
print(planilha.sheetnames)
# Obtendo um sheet para trabalhar
sheet1 = planilha['Sheet1']
sheet1['C3'].value = 'Hashirama' # Alterar o valor da célula
sheet1.cell(row=3, column=3, value='Mark') # Alterar o valor da célula
print(sheet1['C3'].value) # Imprimindo o valor na tela

# Percorrer uma planilha
# for linha in sheet1.iter_rows(min_row=2, max_row=10, min_col=2, max_col=4):
#     print(linha[0].value, linha[1].value, linha[2].value)

for linha in sheet1.iter_cols(min_row=2, min_col=2, max_col=2):
    for cell in linha:
        print(cell.value)
